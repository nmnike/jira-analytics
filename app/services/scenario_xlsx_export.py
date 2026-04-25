"""4-sheet xlsx export for planning scenarios («Бухгалтерия» style).

See docs/superpowers/specs/2026-04-25-scenario-xlsx-export-redesign.md
for layout, colours and contents of each sheet.
"""

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Optional

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from sqlalchemy.orm import Session, joinedload

from app.models import (
    Absence, AppSetting, BacklogItem, Employee, EmployeeTeam,
    MandatoryWorkType, PlanningScenario, ProductionCalendarDay, Role,
    ScenarioAllocation, ScenarioRule,
)
from app.services.resource_base_service import (
    ResourceBaseService, ResourceBase, ResourceSummary,
)


SHEET_NAMES = ["Сводка", "Включено", "Не вошло", "Справочник"]

QUARTER_MONTHS = {1: (1, 2, 3), 2: (4, 5, 6), 3: (7, 8, 9), 4: (10, 11, 12)}

MONTH_LABELS = {
    1: "Янв", 2: "Фев", 3: "Мар", 4: "Апр", 5: "Май", 6: "Июн",
    7: "Июл", 8: "Авг", 9: "Сен", 10: "Окт", 11: "Ноя", 12: "Дек",
}


class _Style:
    """Бухгалтерия palette — dark-grey headers, traffic-light data, no cyan."""

    # Palette
    DARK_HEADER = "1F2937"
    SECTION_BG = "F3F4F6"
    TOTALS_BG = "FAFAFA"
    HEAT_LIGHT = "ECFEFF"
    HEAT_MID = "CFFAFE"
    HEAT_DARK = "67E8F9"
    GREEN_FILL = "DCFCE7"
    GREEN_TEXT = "166534"
    YELLOW_FILL = "FEF3C7"
    YELLOW_TEXT = "92400E"
    RED_FILL = "FEE2E2"
    RED_TEXT = "B91C1C"
    PRI1_TEXT = "B91C1C"
    PRI2_TEXT = "B45309"
    GREY_TEXT = "9CA3AF"
    GREY_FILL = "FAFAFA"
    LINK_BLUE = "1D4ED8"

    # Reusable fills
    HEADER_FILL = PatternFill("solid", fgColor=DARK_HEADER)
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_BG)
    TOTALS_FILL = PatternFill("solid", fgColor=TOTALS_BG)
    GREEN_BG = PatternFill("solid", fgColor=GREEN_FILL)
    YELLOW_BG = PatternFill("solid", fgColor=YELLOW_FILL)
    RED_BG = PatternFill("solid", fgColor=RED_FILL)
    GREY_BG = PatternFill("solid", fgColor=GREY_FILL)

    # Reusable fonts
    STRIP_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="4B5563")
    BOLD_FONT = Font(name="Calibri", size=11, bold=True)
    REGULAR_FONT = Font(name="Calibri", size=11)
    LABEL_FONT = Font(name="Calibri", size=10, color="6B7280")
    ITALIC_GREY_FONT = Font(name="Calibri", size=10, italic=True, color="6B7280")
    PRI1_FONT = Font(name="Calibri", size=11, bold=True, color=PRI1_TEXT)
    PRI2_FONT = Font(name="Calibri", size=11, bold=True, color=PRI2_TEXT)
    LINK_FONT = Font(name="Calibri", size=11, color=LINK_BLUE, underline="single")
    RED_BOLD_FONT = Font(name="Calibri", size=11, bold=True, color=RED_TEXT)
    EXCLUDED_FONT = Font(name="Calibri", size=11, color="6B7280")

    # Borders
    THIN_GREY = Side(style="thin", color="E5E7EB")
    BOTTOM_DARK = Side(style="medium", color=DARK_HEADER)
    TOP_DARK = Side(style="medium", color=DARK_HEADER)
    SECTION_BORDER = Border(bottom=BOTTOM_DARK)
    TOTALS_BORDER = Border(top=TOP_DARK)
    ROW_BORDER = Border(bottom=THIN_GREY)

    # Alignment (wrap_text везде, чтобы длинные значения не вылезали за ячейку)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass
class ScenarioExportContext:
    """Single in-memory snapshot used across all 4 sheets."""

    scenario: PlanningScenario
    allocations: list[ScenarioAllocation]
    resource_summary: ResourceSummary
    resource_base: ResourceBase
    scenario_rules: list[ScenarioRule]
    work_types: list[MandatoryWorkType]
    employees: list[Employee]
    roles_by_code: dict[str, Role]
    absences: list[Absence]
    calendar_by_date: dict[date, float]
    jira_base_url: str
    generated_at: datetime
    period_start: date
    period_end: date  # exclusive


INCLUDED_HEADERS = [
    "Ключ Jira", "Название", "Приоритет", "Заказчик",
    "Аналитик, ч", "Разработка, ч", "QA, ч", "ОПЭ, ч",
    "Итого, ч", "План, ч", "Цели",
]
INCLUDED_WIDTHS = [14, 50, 8, 18, 11, 11, 11, 11, 12, 12, 28]

EXCLUDED_HEADERS = [
    "Ключ Jira", "Название", "Приоритет", "Заказчик",
    "Аналитик, ч", "Разработка, ч", "QA, ч", "ОПЭ, ч",
    "Итого, ч", "Цели",
]
EXCLUDED_WIDTHS = [14, 50, 8, 18, 11, 11, 11, 11, 12, 28]


def _demand_by_role(item: BacklogItem) -> tuple[float, float, float]:
    """Hours by (analyst, dev, qa) accounting for ОПЭ split."""
    ea = item.estimate_analyst_hours or 0.0
    ed = item.estimate_dev_hours or 0.0
    eq = item.estimate_qa_hours or 0.0
    eo = item.estimate_opo_hours or 0.0
    r = item.opo_analyst_ratio if item.opo_analyst_ratio is not None else 0.5
    return ea + eo * r, ed + eo * (1.0 - r), eq


def _initiative_row_mid(alloc: ScenarioAllocation, *, included: bool) -> list:
    """Row values for Mid-11 (included) or Mid-10 (excluded) sheets."""
    item = alloc.backlog_item
    issue = getattr(item, "issue", None)
    key = issue.key if issue else ""
    analyst, dev, qa = _demand_by_role(item)
    opo = item.estimate_opo_hours or 0.0
    total = round(analyst + dev + qa + opo, 1)
    goals = (issue.goals or "") if issue else ""
    base = [
        key,
        item.title,
        item.priority,
        item.customer or "",
        round(analyst, 1),
        round(dev, 1),
        round(qa, 1),
        round(opo, 1),
        total,
    ]
    if included:
        base.append(round(alloc.planned_hours or 0.0, 1))
    base.append(goals)
    return base


def _write_title_strip(ws, text: str, *, columns: int = 8) -> None:
    """Dark-grey header strip merged across N columns, white bold text."""
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = _Style.STRIP_FONT
    cell.fill = _Style.HEADER_FILL
    cell.alignment = _Style.LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws.row_dimensions[1].height = 22


def _write_subtitle(ws, text: str, *, row: int = 2, columns: int = 8) -> None:
    """Light-grey subtitle row below the title strip."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _Style.LABEL_FONT
    cell.fill = _Style.SECTION_FILL
    cell.alignment = _Style.LEFT
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=columns,
    )


def _write_section_header(ws, row: int, text: str, *, columns: int = 8) -> None:
    """`СЕКЦИЯ` row with grey fill + dark bottom border, merged."""
    cell = ws.cell(row=row, column=1, value=text.upper())
    cell.font = _Style.SECTION_FONT
    cell.fill = _Style.SECTION_FILL
    cell.alignment = _Style.LEFT
    cell.border = _Style.SECTION_BORDER
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=columns,
    )


def _stamp_row_borders(
    ws, ranges: list[tuple[int, int, int]],
) -> None:
    """Тонкий нижний border + wrap_text на каждую ячейку data-строки.

    `ranges` — список `(start_row, end_row, columns)`. Сохраняет уже
    выставленные left/right/top и horizontal-alignment, перезаписывает
    только bottom-border и включает wrap_text.
    """
    for start, end, cols in ranges:
        if end < start:
            continue
        for r in range(start, end + 1):
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                # bottom border
                existing = cell.border
                cell.border = Border(
                    left=existing.left,
                    right=existing.right,
                    top=existing.top,
                    bottom=_Style.THIN_GREY,
                )
                # wrap_text — сохраняем horizontal/vertical если заданы
                a = cell.alignment
                cell.alignment = Alignment(
                    horizontal=(a.horizontal if a is not None else None),
                    vertical=(a.vertical if (a is not None and a.vertical) else "center"),
                    wrap_text=True,
                )


ANALYST_SUBSTITUTE_ROLES = {"project_manager", "consultant", "RP"}


def _planned_hours_by_role(ctx: ScenarioExportContext) -> dict[str, float]:
    """Запланированные часы по ролям с учётом исполнителя.

    Программист и тестировщик всегда попадают в свои роли (`dev` / `qa`).
    Аналитический объём попадает в `analyst`, **кроме** случая, когда исполнитель —
    Руководитель проектов или Консультант: тогда аналитический объём списывается
    на роль исполнителя. Логика повторяет `demandByAssigneeRole` на фронте.
    """
    out: dict[str, float] = {}
    for a in ctx.allocations:
        if not a.included_flag:
            continue
        analyst, dev, qa = _demand_by_role(a.backlog_item)
        total_est = analyst + dev + qa
        if total_est <= 0:
            continue
        p = a.planned_hours or 0.0
        analyst_target = "analyst"
        assignee = getattr(a.backlog_item, "assignee", None)
        assignee_role = assignee.role if assignee is not None else None
        if assignee_role and assignee_role in ANALYST_SUBSTITUTE_ROLES:
            analyst_target = assignee_role
        out[analyst_target] = out.get(analyst_target, 0.0) + p * analyst / total_est
        out["dev"] = out.get("dev", 0.0) + p * dev / total_est
        out["qa"] = out.get("qa", 0.0) + p * qa / total_est
    return out


def _absence_hours_in_period(absence: Absence, ctx: ScenarioExportContext) -> float:
    """Часов отсутствия, попадающих в квартал. Использует hours_total если задано."""
    if absence.hours_total is not None:
        return float(absence.hours_total)
    h = 0.0
    cur = max(absence.start_date, ctx.period_start)
    while cur <= absence.end_date and cur < ctx.period_end:
        norm = ctx.calendar_by_date.get(
            cur, 8.0 if cur.weekday() < 5 else 0.0,
        )
        if norm > 0:
            h += norm
        cur = cur + timedelta(days=1)
    return round(h, 1)


def _absence_days_in_period(absence: Absence, ctx: ScenarioExportContext) -> int:
    """Рабочих дней отсутствия, попадающих в квартал."""
    cnt = 0
    cur = max(absence.start_date, ctx.period_start)
    while cur <= absence.end_date and cur < ctx.period_end:
        norm = ctx.calendar_by_date.get(
            cur, 8.0 if cur.weekday() < 5 else 0.0,
        )
        if norm > 0:
            cnt += 1
        cur = cur + timedelta(days=1)
    return cnt


def _per_role_per_month(ctx: ScenarioExportContext) -> dict[tuple[str, int], float]:
    """role_code -> month -> доступные часы (после вычета отсутствий)."""
    out: dict[tuple[str, int], float] = {}
    for emp in ctx.resource_base.employees:
        if not emp.role:
            continue
        for d in emp.days:
            key = (emp.role, d.date.month)
            out[key] = out.get(key, 0.0) + d.hours
    # External QA override: spread evenly across the quarter's 3 months.
    if ctx.scenario.external_qa_hours is not None:
        q = int(str(ctx.scenario.quarter or "Q1").replace("Q", ""))
        months = QUARTER_MONTHS[q]
        ext = float(ctx.scenario.external_qa_hours)
        per_month = round(ext / 3.0, 1)
        out[("qa", months[0])] = per_month
        out[("qa", months[1])] = per_month
        # Last month gets the remainder so the total matches exactly.
        out[("qa", months[2])] = round(ext - per_month * 2, 1)
    return out


class ScenarioXlsxExporter:
    """Build a 4-sheet xlsx for a planning scenario.

    Public API:
        ScenarioXlsxExporter(db, scenario_id).build() -> bytes
    """

    def __init__(self, db: Session, scenario_id: str) -> None:
        self.db = db
        self.scenario_id = scenario_id

    def build(self) -> bytes:
        ctx = self._load()
        wb = Workbook()
        wb.remove(wb.active)
        for name in SHEET_NAMES:
            wb.create_sheet(name)

        self._sheet_summary(wb["Сводка"], ctx)
        self._sheet_included(wb["Включено"], ctx)
        self._sheet_excluded(wb["Не вошло"], ctx)
        self._sheet_reference(wb["Справочник"], ctx)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # === Loading ===

    def _load(self) -> ScenarioExportContext:
        scenario = self.db.get(PlanningScenario, self.scenario_id)
        if scenario is None:
            raise ValueError(f"Scenario {self.scenario_id} not found")

        q = int(str(scenario.quarter or "Q1").replace("Q", ""))
        year = scenario.year or datetime.utcnow().year
        months = QUARTER_MONTHS[q]
        period_start = date(year, months[0], 1)
        last_m = months[-1]
        next_year = year + 1 if last_m == 12 else year
        next_month = 1 if last_m == 12 else last_m + 1
        period_end = date(next_year, next_month, 1)

        allocations = (
            self.db.query(ScenarioAllocation)
            .options(
                joinedload(ScenarioAllocation.backlog_item)
                    .joinedload(BacklogItem.issue),
                joinedload(ScenarioAllocation.backlog_item)
                    .joinedload(BacklogItem.project),
                joinedload(ScenarioAllocation.backlog_item)
                    .joinedload(BacklogItem.assignee),
            )
            .filter(ScenarioAllocation.scenario_id == self.scenario_id)
            .all()
        )

        rb_service = ResourceBaseService(self.db)
        resource_base = rb_service.compute(scenario)
        resource_summary = rb_service.compute_summary(scenario)

        scenario_rules = (
            self.db.query(ScenarioRule)
            .filter(ScenarioRule.scenario_id == self.scenario_id)
            .all()
        )

        work_types = (
            self.db.query(MandatoryWorkType)
            .filter(MandatoryWorkType.is_active == True)  # noqa: E712
            .order_by(MandatoryWorkType.sort_order.asc())
            .all()
        )

        emp_ids = [
            r[0]
            for r in self.db.query(EmployeeTeam.employee_id)
                .filter(EmployeeTeam.team == scenario.team)
                .all()
        ]
        employees = (
            self.db.query(Employee)
            .filter(Employee.id.in_(emp_ids), Employee.is_active == True)  # noqa: E712
            .all()
        ) if emp_ids else []

        roles_by_code = {r.code: r for r in self.db.query(Role).all()}

        absences = (
            self.db.query(Absence)
            .options(joinedload(Absence.reason))
            .filter(
                Absence.employee_id.in_(emp_ids) if emp_ids else False,  # type: ignore[arg-type]
                Absence.start_date < period_end,
                Absence.end_date >= period_start,
            )
            .all()
        ) if emp_ids else []

        calendar_by_date = {
            row.date: float(row.hours)
            for row in self.db.query(ProductionCalendarDay).filter(
                ProductionCalendarDay.date >= period_start,
                ProductionCalendarDay.date < period_end,
            ).all()
        }

        base_url_setting = self.db.get(AppSetting, "jira_base_url")
        jira_base_url = (base_url_setting.value if base_url_setting else "") or ""

        return ScenarioExportContext(
            scenario=scenario,
            allocations=allocations,
            resource_summary=resource_summary,
            resource_base=resource_base,
            scenario_rules=scenario_rules,
            work_types=work_types,
            employees=employees,
            roles_by_code=roles_by_code,
            absences=absences,
            calendar_by_date=calendar_by_date,
            jira_base_url=jira_base_url,
            generated_at=datetime.utcnow(),
            period_start=period_start,
            period_end=period_end,
        )

    # === Sheets ===

    def _sheet_summary(self, ws, ctx: ScenarioExportContext) -> None:
        ws.sheet_view.showGridLines = False
        summary = ctx.resource_summary

        # --- Title strip + subtitle ---
        status_label = "утверждён" if ctx.scenario.status == "approved" else "черновик"
        title = f"Сценарий: {ctx.scenario.name} — {status_label}"
        _write_title_strip(ws, title, columns=8)

        subtitle = (
            f"{ctx.scenario.quarter or ''} "
            f"{ctx.scenario.year or ''} · команда «{ctx.scenario.team or '—'}» · "
            f"сформировано {ctx.generated_at:%d.%m.%Y}"
        ).strip()
        _write_subtitle(ws, subtitle, row=2, columns=8)

        # --- Section 1: Сводка metrics 2×8 ---
        _write_section_header(ws, row=4, text="Сводка", columns=8)

        # Compute metrics
        total_capacity = round(summary.available_total, 1)
        planned_by_role = _planned_hours_by_role(ctx)
        total_planned = round(sum(planned_by_role.values()), 1)
        leftover = round(max(0.0, total_capacity - total_planned), 1)
        included_n = sum(1 for a in ctx.allocations if a.included_flag)
        excluded_n = sum(1 for a in ctx.allocations if not a.included_flag)
        qa_avail = summary.available_by_role.get("qa", 0.0)
        qa_planned = planned_by_role.get("qa", 0.0)
        qa_deficit = round(min(0.0, qa_avail - qa_planned), 1)
        usage_pct = (
            round(total_planned / total_capacity, 4) if total_capacity > 0 else 0.0
        )
        absences_total_hours = round(
            sum(_absence_hours_in_period(a, ctx) for a in ctx.absences), 1,
        )

        # Row 5
        pairs_5 = [
            ("Ёмкость, ч", total_capacity, "#,##0"),
            ("План, ч", total_planned, "#,##0"),
            ("Остаток, ч", leftover, "#,##0"),
            ("Использование", usage_pct, "0%"),
        ]
        for i, (label, value, fmt) in enumerate(pairs_5):
            col = 1 + i * 2
            lc = ws.cell(row=5, column=col, value=label)
            lc.font = _Style.LABEL_FONT
            lc.alignment = _Style.LEFT
            v = ws.cell(row=5, column=col + 1, value=value)
            v.font = _Style.BOLD_FONT
            v.alignment = _Style.RIGHT
            v.number_format = fmt
        # Usage % conditional fill (col 8)
        if total_capacity > 0:
            v = ws.cell(row=5, column=8)
            if usage_pct < 0.8:
                v.fill = _Style.GREEN_BG
            elif usage_pct <= 1.10:
                v.fill = _Style.YELLOW_BG
            else:
                v.fill = _Style.RED_BG
                v.font = _Style.RED_BOLD_FONT

        # Row 6
        pairs_6 = [
            ("Включено, шт.", included_n, "#,##0"),
            ("Не вошло, шт.", excluded_n, "#,##0"),
            ("QA дефицит, ч", qa_deficit, "#,##0"),
            ("Отсутствия, ч", absences_total_hours, "#,##0"),
        ]
        for i, (label, value, fmt) in enumerate(pairs_6):
            col = 1 + i * 2
            lc = ws.cell(row=6, column=col, value=label)
            lc.font = _Style.LABEL_FONT
            lc.alignment = _Style.LEFT
            v = ws.cell(row=6, column=col + 1, value=value)
            v.font = _Style.BOLD_FONT
            v.alignment = _Style.RIGHT
            v.number_format = fmt
        # QA дефицит red bold if negative
        if qa_deficit < 0:
            ws.cell(row=6, column=6).font = _Style.RED_BOLD_FONT

        # --- Section 2: По ролям × месяцам ---
        section2_row = 8
        _write_section_header(ws, row=section2_row, text="По ролям × месяцам", columns=8)

        months = QUARTER_MONTHS[int(str(ctx.scenario.quarter or "Q1").replace("Q", ""))]
        month_labels = [MONTH_LABELS[m] for m in months]

        header_row = section2_row + 1
        headers_2 = ["Роль", *month_labels, "Σ Ёмкость", "План", "Остаток", "% исп."]
        for c_idx, h in enumerate(headers_2, start=1):
            c = ws.cell(row=header_row, column=c_idx, value=h)
            c.font = _Style.HEADER_FONT
            c.fill = _Style.HEADER_FILL
            c.alignment = _Style.CENTER if c_idx > 1 else _Style.LEFT

        role_month = _per_role_per_month(ctx)
        planned_by_role = _planned_hours_by_role(ctx)

        total_per_month = [0.0, 0.0, 0.0]
        total_capacity_sum = 0.0
        total_planned_sum = 0.0

        r_idx = header_row + 1
        for role in summary.roles:
            role_label = (
                ctx.roles_by_code[role].label if role in ctx.roles_by_code else role
            )
            ws.cell(row=r_idx, column=1, value=role_label).font = _Style.BOLD_FONT

            m_sum = 0.0
            for i, m in enumerate(months):
                v = round(role_month.get((role, m), 0.0), 1)
                cell = ws.cell(row=r_idx, column=2 + i, value=v)
                cell.alignment = _Style.RIGHT
                cell.number_format = "#,##0"
                m_sum += v
                total_per_month[i] += v

            capacity = summary.available_by_role.get(role, 0.0)
            ws.cell(row=r_idx, column=5, value=round(capacity, 1)).number_format = "#,##0"
            ws.cell(row=r_idx, column=5).font = _Style.BOLD_FONT
            ws.cell(row=r_idx, column=5).alignment = _Style.RIGHT
            total_capacity_sum += capacity

            planned = round(planned_by_role.get(role, 0.0), 1)
            ws.cell(row=r_idx, column=6, value=planned).number_format = "#,##0"
            ws.cell(row=r_idx, column=6).alignment = _Style.RIGHT
            total_planned_sum += planned

            leftover_r = round(capacity - planned, 1)
            leftover_cell = ws.cell(row=r_idx, column=7, value=leftover_r)
            leftover_cell.number_format = "#,##0"
            leftover_cell.alignment = _Style.RIGHT
            if leftover_r < 0:
                leftover_cell.font = _Style.RED_BOLD_FONT

            usage = (planned / capacity) if capacity > 0 else 0.0
            usage_cell = ws.cell(row=r_idx, column=8, value=round(usage, 4))
            usage_cell.number_format = "0%"
            usage_cell.alignment = _Style.RIGHT
            if usage < 0.8:
                usage_cell.fill = _Style.GREEN_BG
            elif usage <= 1.10:
                usage_cell.fill = _Style.YELLOW_BG
            else:
                usage_cell.fill = _Style.RED_BG
                usage_cell.font = _Style.RED_BOLD_FONT

            r_idx += 1

        # ИТОГО row
        total_row_idx = r_idx
        total_cell = ws.cell(row=total_row_idx, column=1, value="ИТОГО")
        total_cell.font = _Style.BOLD_FONT
        total_cell.fill = _Style.TOTALS_FILL
        total_cell.border = _Style.TOTALS_BORDER
        for i, v in enumerate(total_per_month):
            c = ws.cell(row=total_row_idx, column=2 + i, value=round(v, 1))
            c.font = _Style.BOLD_FONT
            c.fill = _Style.TOTALS_FILL
            c.border = _Style.TOTALS_BORDER
            c.number_format = "#,##0"
            c.alignment = _Style.RIGHT
        for c_idx, value in [
            (5, round(total_capacity_sum, 1)),
            (6, round(total_planned_sum, 1)),
            (7, round(total_capacity_sum - total_planned_sum, 1)),
        ]:
            c = ws.cell(row=total_row_idx, column=c_idx, value=value)
            c.font = _Style.BOLD_FONT
            c.fill = _Style.TOTALS_FILL
            c.border = _Style.TOTALS_BORDER
            c.number_format = "#,##0"
            c.alignment = _Style.RIGHT
        total_usage = (
            total_planned_sum / total_capacity_sum if total_capacity_sum > 0 else 0.0
        )
        c = ws.cell(row=total_row_idx, column=8, value=round(total_usage, 4))
        c.font = _Style.BOLD_FONT
        c.fill = _Style.TOTALS_FILL
        c.border = _Style.TOTALS_BORDER
        c.number_format = "0%"
        c.alignment = _Style.RIGHT

        # --- Section 3: По сотрудникам ---
        section3_row = total_row_idx + 2
        _write_section_header(ws, row=section3_row, text="По сотрудникам", columns=8)

        emp_header_row = section3_row + 1
        emp_headers = [
            "Сотрудник", "Роль", "Норма, ч", "Отсутствия, ч",
            "Доступно, ч", "Дней отсутствия",
        ]
        for c_idx, h in enumerate(emp_headers, start=1):
            c = ws.cell(row=emp_header_row, column=c_idx, value=h)
            c.font = _Style.HEADER_FONT
            c.fill = _Style.HEADER_FILL
            c.alignment = _Style.CENTER if c_idx > 1 else _Style.LEFT

        base_by_id = {e.employee_id: e for e in ctx.resource_base.employees}

        emp_rows = []
        for emp in ctx.employees:
            role_label = (
                ctx.roles_by_code[emp.role].label
                if emp.role and emp.role in ctx.roles_by_code else (emp.role or "—")
            )
            base = base_by_id.get(emp.id)
            cal_gross = 0.0
            cur = ctx.period_start
            while cur < ctx.period_end:
                cal_gross += ctx.calendar_by_date.get(
                    cur, 8.0 if cur.weekday() < 5 else 0.0,
                )
                cur = cur + timedelta(days=1)

            available = base.total_hours if base else 0.0
            emp_abs = [a for a in ctx.absences if a.employee_id == emp.id]
            abs_days = 0
            cur = ctx.period_start
            while cur < ctx.period_end:
                norm = ctx.calendar_by_date.get(
                    cur, 8.0 if cur.weekday() < 5 else 0.0,
                )
                if norm > 0:
                    if any(a.start_date <= cur <= a.end_date for a in emp_abs):
                        abs_days += 1
                cur = cur + timedelta(days=1)
            absence_hours = round(cal_gross - available, 1)

            emp_rows.append({
                "name": emp.display_name,
                "role": role_label,
                "norm": round(cal_gross, 1),
                "absence": absence_hours,
                "available": round(available, 1),
                "abs_days": abs_days,
            })

        emp_rows.sort(key=lambda r: (r["role"], r["name"]))

        r_idx = emp_header_row + 1
        for row in emp_rows:
            ws.cell(row=r_idx, column=1, value=row["name"])
            ws.cell(row=r_idx, column=2, value=row["role"])
            for c_idx, key, fmt in [
                (3, "norm", "#,##0"),
                (4, "absence", "#,##0"),
                (5, "available", "#,##0"),
            ]:
                c = ws.cell(row=r_idx, column=c_idx, value=row[key])
                c.number_format = fmt
                c.alignment = _Style.RIGHT
            ws.cell(row=r_idx, column=5).font = _Style.BOLD_FONT
            days_cell = ws.cell(row=r_idx, column=6, value=row["abs_days"])
            days_cell.alignment = _Style.RIGHT
            abs_days_val = row["abs_days"]  # type: ignore[assignment]
            if 6 <= abs_days_val <= 15:  # type: ignore[operator]
                days_cell.fill = _Style.YELLOW_BG
            elif abs_days_val > 15:  # type: ignore[operator]
                days_cell.fill = _Style.RED_BG
            r_idx += 1

        # Тонкая линия снизу каждой data-строки в обеих секциях
        _stamp_row_borders(ws, [
            (header_row + 1, total_row_idx - 1, 8),  # «По ролям × месяцам»
            (emp_header_row + 1, r_idx - 1, 8),      # «По сотрудникам»
        ])

        # --- Column widths ---
        widths = [26, 18, 12, 12, 12, 12, 12, 12]
        for c_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        ws.freeze_panes = "A4"

    def _sheet_included(self, ws, ctx: ScenarioExportContext) -> None:
        from openpyxl.formatting.rule import ColorScaleRule  # type: ignore[import-untyped]

        ws.sheet_view.showGridLines = False

        rows = sorted(
            [a for a in ctx.allocations if a.included_flag],
            key=lambda a: (
                a.backlog_item.priority is None,
                a.backlog_item.priority if a.backlog_item.priority is not None else 0,
                a.backlog_item.title,
            ),
        )

        # Title strip
        status = "утверждён" if ctx.scenario.status == "approved" else "черновик"
        title = (
            f"Сценарий: {ctx.scenario.name} — {status} · "
            f"Включено ({len(rows)} задач)"
        )
        _write_title_strip(ws, title, columns=len(INCLUDED_HEADERS))

        # Header row at row 2
        for c_idx, h in enumerate(INCLUDED_HEADERS, start=1):
            c = ws.cell(row=2, column=c_idx, value=h)
            c.font = _Style.HEADER_FONT
            c.fill = _Style.HEADER_FILL
            c.alignment = _Style.CENTER
        ws.row_dimensions[2].height = 22

        # Data rows
        for r_idx, alloc in enumerate(rows, start=3):
            values = _initiative_row_mid(alloc, included=True)
            for c_idx, val in enumerate(values, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                if c_idx in (5, 6, 7, 8):
                    c.number_format = "#,##0.#"
                    c.alignment = _Style.RIGHT
                elif c_idx == 3:
                    c.alignment = _Style.CENTER
                    if val == 1:
                        c.font = _Style.PRI1_FONT
                    elif val == 2:
                        c.font = _Style.PRI2_FONT
                elif c_idx == 9:
                    c.font = _Style.BOLD_FONT
                    c.number_format = "#,##0.#"
                    c.alignment = _Style.RIGHT
                    c.fill = PatternFill("solid", fgColor="EFF6FF")
                elif c_idx == 10:
                    c.font = _Style.BOLD_FONT
                    c.number_format = "#,##0.#"
                    c.alignment = _Style.RIGHT

            # Hyperlink on key column
            key = values[0]
            if key and ctx.jira_base_url:
                link = f"{ctx.jira_base_url.rstrip('/')}/browse/{key}"
                ws.cell(row=r_idx, column=1).hyperlink = link
                ws.cell(row=r_idx, column=1).font = _Style.LINK_FONT

        # Totals row
        total_row_idx = len(rows) + 3
        total_label = ws.cell(row=total_row_idx, column=1, value=f"Σ ИТОГО ({len(rows)} задач)")
        total_label.font = _Style.HEADER_FONT
        total_label.fill = _Style.HEADER_FILL
        ws.merge_cells(
            start_row=total_row_idx, start_column=1,
            end_row=total_row_idx, end_column=4,
        )
        sum_cols = [5, 6, 7, 8, 9, 10]
        for c_idx in sum_cols:
            if rows:
                total = sum(_initiative_row_mid(a, included=True)[c_idx - 1] for a in rows)
            else:
                total = 0
            c = ws.cell(row=total_row_idx, column=c_idx, value=round(total, 1))
            c.font = _Style.HEADER_FONT
            c.fill = _Style.HEADER_FILL
            c.number_format = "#,##0.#"
            c.alignment = _Style.RIGHT
        # Empty cell for "Цели" column to keep the strip continuous
        c = ws.cell(row=total_row_idx, column=11, value="")
        c.fill = _Style.HEADER_FILL

        # Тонкая линия снизу каждой строки данных
        _stamp_row_borders(ws, [(3, total_row_idx - 1, len(INCLUDED_HEADERS))])

        # Heatmap on hours columns 5-8 (only if there are rows)
        if rows:
            for c_idx in (5, 6, 7, 8):
                col = get_column_letter(c_idx)
                rng = f"{col}3:{col}{total_row_idx - 1}"
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min", start_color=_Style.HEAT_LIGHT,
                        end_type="max", end_color=_Style.HEAT_DARK,
                    ),
                )

        # Column widths
        for c_idx, w in enumerate(INCLUDED_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        ws.freeze_panes = "A3"
        if rows:
            ws.auto_filter.ref = f"A2:{get_column_letter(len(INCLUDED_HEADERS))}{total_row_idx}"
        else:
            ws.auto_filter.ref = f"A2:{get_column_letter(len(INCLUDED_HEADERS))}2"

    def _sheet_excluded(self, ws, ctx: ScenarioExportContext) -> None:
        from openpyxl.formatting.rule import ColorScaleRule  # type: ignore[import-untyped]

        ws.sheet_view.showGridLines = False

        rows = sorted(
            [a for a in ctx.allocations if not a.included_flag],
            key=lambda a: (
                a.backlog_item.priority is None,
                a.backlog_item.priority if a.backlog_item.priority is not None else 0,
                a.backlog_item.title,
            ),
        )

        title = f"Сценарий: {ctx.scenario.name} · Не вошло ({len(rows)} задач)"
        _write_title_strip(ws, title, columns=len(EXCLUDED_HEADERS))

        for c_idx, h in enumerate(EXCLUDED_HEADERS, start=1):
            c = ws.cell(row=2, column=c_idx, value=h)
            c.font = _Style.HEADER_FONT
            c.fill = _Style.HEADER_FILL
            c.alignment = _Style.CENTER
        ws.row_dimensions[2].height = 22

        for r_idx, alloc in enumerate(rows, start=3):
            values = _initiative_row_mid(alloc, included=False)
            for c_idx, val in enumerate(values, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.fill = _Style.GREY_BG
                if c_idx in (5, 6, 7, 8):
                    c.number_format = "#,##0.#"
                    c.alignment = _Style.RIGHT
                elif c_idx == 3:
                    c.alignment = _Style.CENTER
                elif c_idx == 9:
                    c.font = _Style.BOLD_FONT
                    c.number_format = "#,##0.#"
                    c.alignment = _Style.RIGHT
                elif c_idx in (2, 10):
                    c.font = _Style.EXCLUDED_FONT
            # Hyperlink on key column
            key = values[0]
            if key and ctx.jira_base_url:
                link = f"{ctx.jira_base_url.rstrip('/')}/browse/{key}"
                ws.cell(row=r_idx, column=1).hyperlink = link
                ws.cell(row=r_idx, column=1).font = _Style.LINK_FONT

        # Totals row
        total_row_idx = len(rows) + 3
        total_label = ws.cell(
            row=total_row_idx, column=1,
            value=f"Σ ИТОГО ({len(rows)} задач не вошло)",
        )
        total_label.font = _Style.HEADER_FONT
        total_label.fill = _Style.HEADER_FILL
        ws.merge_cells(
            start_row=total_row_idx, start_column=1,
            end_row=total_row_idx, end_column=4,
        )
        sum_cols = [5, 6, 7, 8, 9]
        for c_idx in sum_cols:
            if rows:
                total = sum(_initiative_row_mid(a, included=False)[c_idx - 1] for a in rows)
            else:
                total = 0
            c = ws.cell(row=total_row_idx, column=c_idx, value=round(total, 1))
            c.font = _Style.HEADER_FONT
            c.fill = _Style.HEADER_FILL
            c.number_format = "#,##0.#"
            c.alignment = _Style.RIGHT
        c = ws.cell(row=total_row_idx, column=10, value="")
        c.fill = _Style.HEADER_FILL

        # Тонкая линия снизу каждой строки данных
        _stamp_row_borders(ws, [(3, total_row_idx - 1, len(EXCLUDED_HEADERS))])

        # Heatmap on hours
        if rows:
            for c_idx in (5, 6, 7, 8):
                col = get_column_letter(c_idx)
                rng = f"{col}3:{col}{total_row_idx - 1}"
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min", start_color=_Style.HEAT_LIGHT,
                        end_type="max", end_color=_Style.HEAT_DARK,
                    ),
                )

        for c_idx, w in enumerate(EXCLUDED_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        ws.freeze_panes = "A3"
        if rows:
            ws.auto_filter.ref = f"A2:{get_column_letter(len(EXCLUDED_HEADERS))}{total_row_idx}"
        else:
            ws.auto_filter.ref = f"A2:{get_column_letter(len(EXCLUDED_HEADERS))}2"

    def _sheet_reference(self, ws, ctx: ScenarioExportContext) -> None:
        ws.sheet_view.showGridLines = False
        summary = ctx.resource_summary
        sub_wts = [w for w in ctx.work_types if w.subtracts_from_pool]

        # --- Title strip ---
        title = (
            f"Справочник · {ctx.scenario.quarter or ''} {ctx.scenario.year or ''} · "
            f"{ctx.scenario.team or '—'}"
        )
        _write_title_strip(ws, title, columns=8)
        ws.cell(row=2, column=1).value = ""  # blank row 2 for spacing

        # --- Section 1: Правила распределения часов ---
        sec1_row = 4
        _write_section_header(
            ws, row=sec1_row,
            text="1 · Правила распределения часов на обязательные работы",
            columns=8,
        )
        hint = ws.cell(row=sec1_row + 1, column=1, value=(
            "Сколько процентов нормы каждой роли резервируется на обязательные виды работ. "
            "Остаток уходит на инициативы."
        ))
        hint.font = _Style.ITALIC_GREY_FONT
        hint.alignment = _Style.LEFT
        ws.merge_cells(
            start_row=sec1_row + 1, start_column=1,
            end_row=sec1_row + 1, end_column=8,
        )

        # Matrix header
        matrix_header_row = sec1_row + 2
        ws.cell(row=matrix_header_row, column=1, value="Роль").font = _Style.HEADER_FONT
        ws.cell(row=matrix_header_row, column=1).fill = _Style.HEADER_FILL
        ws.cell(row=matrix_header_row, column=1).alignment = _Style.LEFT
        for c_idx, wt in enumerate(sub_wts, start=2):
            c = ws.cell(row=matrix_header_row, column=c_idx, value=wt.label)
            c.font = _Style.HEADER_FONT
            c.fill = _Style.HEADER_FILL
            c.alignment = _Style.CENTER
        sum_col_idx = len(sub_wts) + 2
        c = ws.cell(row=matrix_header_row, column=sum_col_idx, value="Σ обязат.")
        c.font = _Style.HEADER_FONT
        c.fill = _Style.HEADER_FILL
        c.alignment = _Style.CENTER

        # Rule lookup
        rule_lookup: dict[tuple[str, Optional[str]], float] = {}
        for r in ctx.scenario_rules:
            key = (r.work_type_id, r.role)
            rule_lookup[key] = rule_lookup.get(key, 0.0) + r.percent_of_norm

        r_idx = matrix_header_row + 1
        roles_to_render = list(summary.roles)
        has_null_rule = any(rule.role is None for rule in ctx.scenario_rules)
        if has_null_rule:
            roles_to_render.append(None)  # type: ignore[arg-type]

        for role in roles_to_render:
            role_label = (
                ctx.roles_by_code[role].label
                if role and role in ctx.roles_by_code else (role or "Все роли")
            )
            ws.cell(row=r_idx, column=1, value=role_label).font = _Style.BOLD_FONT
            sum_pct = 0.0
            gross = summary.gross_by_role.get(role, 0.0) if role else 0.0
            for c_idx, wt in enumerate(sub_wts, start=2):
                pct = rule_lookup.get((wt.id, role))
                if pct is None:
                    cell = ws.cell(row=r_idx, column=c_idx, value=None)
                    cell.fill = _Style.GREY_BG
                    continue
                hours = round(gross * pct / 100.0, 1) if role else None
                text = f"{pct:.0f}%" + (f" · {hours:.0f} ч" if hours is not None else "")
                cell = ws.cell(row=r_idx, column=c_idx, value=text)
                cell.alignment = _Style.RIGHT
                sum_pct += pct
                if pct >= 50:
                    cell.fill = PatternFill("solid", fgColor=_Style.HEAT_DARK)
                elif pct >= 25:
                    cell.fill = PatternFill("solid", fgColor=_Style.HEAT_MID)
                elif pct >= 10:
                    cell.fill = PatternFill("solid", fgColor=_Style.HEAT_LIGHT)
            sum_cell = ws.cell(row=r_idx, column=sum_col_idx, value=f"{sum_pct:.0f}%")
            sum_cell.alignment = _Style.RIGHT
            sum_cell.font = _Style.BOLD_FONT
            if sum_pct > 100:
                sum_cell.font = _Style.RED_BOLD_FONT
            r_idx += 1

        # Тонкая линия снизу каждой строки матрицы правил
        _stamp_row_borders(ws, [(matrix_header_row + 1, r_idx - 1, sum_col_idx)])

        # --- Section 2: Внешний QA-лимит ---
        sec2_row = r_idx + 1
        _write_section_header(ws, row=sec2_row, text="2 · Внешний QA-лимит", columns=8)
        qa_label_cell = ws.cell(
            row=sec2_row + 1, column=1,
            value="Часы внешнего QA, фиксированный лимит на квартал",
        )
        qa_label_cell.alignment = _Style.LEFT
        qa_value = ctx.scenario.external_qa_hours
        if qa_value is None:
            v = ws.cell(row=sec2_row + 1, column=2, value="не задан")
            v.font = _Style.ITALIC_GREY_FONT
        else:
            v = ws.cell(row=sec2_row + 1, column=2, value=round(float(qa_value), 1))
            v.font = _Style.BOLD_FONT
            v.number_format = "#,##0"
        v.alignment = _Style.RIGHT
        note = ws.cell(
            row=sec2_row + 2, column=1,
            value="Замещает суммарную ёмкость штатных QA, если задано",
        )
        note.font = _Style.ITALIC_GREY_FONT

        # --- Section 3: Отсутствия команды ---
        sec3_row = sec2_row + 4
        _write_section_header(
            ws, row=sec3_row, text="3 · Отсутствия команды в квартале", columns=8,
        )

        emp_by_id = {e.id: e for e in ctx.employees}
        abs_rows = []
        for a in ctx.absences:
            emp = emp_by_id.get(a.employee_id)
            if emp is None:
                continue
            role_label = (
                ctx.roles_by_code[emp.role].label
                if emp.role and emp.role in ctx.roles_by_code else (emp.role or "—")
            )
            abs_rows.append({
                "name": emp.display_name,
                "role": role_label,
                "reason": a.reason.label if a.reason else "—",
                "color": a.reason.color if a.reason else None,
                "kind": "Плановая" if (a.reason and a.reason.is_planned) else "Внеплановая",
                "is_planned": bool(a.reason and a.reason.is_planned),
                "start": a.start_date,
                "end": a.end_date,
                "days": _absence_days_in_period(a, ctx),
                "hours": _absence_hours_in_period(a, ctx),
            })
        abs_rows.sort(key=lambda r: (r["name"], r["start"]))

        if not abs_rows:
            msg_cell = ws.cell(
                row=sec3_row + 1, column=1, value="Отсутствий в квартале нет",
            )
            msg_cell.font = _Style.ITALIC_GREY_FONT
        else:
            # Header
            abs_header_row = sec3_row + 1
            abs_headers = ["Сотрудник", "Роль", "Причина", "Тип",
                           "Начало", "Конец", "Дней", "Часов"]
            for c_idx, h in enumerate(abs_headers, start=1):
                c = ws.cell(row=abs_header_row, column=c_idx, value=h)
                c.font = _Style.HEADER_FONT
                c.fill = _Style.HEADER_FILL
                c.alignment = _Style.CENTER if c_idx > 1 else _Style.LEFT

            r_idx = abs_header_row + 1
            for row in abs_rows:
                ws.cell(row=r_idx, column=1, value=row["name"])
                ws.cell(row=r_idx, column=2, value=row["role"])
                reason_cell = ws.cell(row=r_idx, column=3, value=row["reason"])
                if row["color"]:
                    color_hex = str(row["color"]).lstrip("#").upper()
                    reason_cell.fill = PatternFill("solid", fgColor=color_hex)
                    reason_cell.font = Font(
                        name="Calibri", size=11, bold=True, color="FFFFFF",
                    )
                    reason_cell.alignment = _Style.CENTER
                kind_cell = ws.cell(row=r_idx, column=4, value=row["kind"])
                kind_cell.fill = _Style.GREEN_BG if row["is_planned"] else _Style.RED_BG
                kind_cell.alignment = _Style.CENTER
                ws.cell(row=r_idx, column=5, value=row["start"]).number_format = "DD.MM.YYYY"
                ws.cell(row=r_idx, column=6, value=row["end"]).number_format = "DD.MM.YYYY"
                d = ws.cell(row=r_idx, column=7, value=row["days"])
                d.number_format = "#,##0"
                d.alignment = _Style.RIGHT
                h_cell = ws.cell(row=r_idx, column=8, value=row["hours"])
                h_cell.number_format = "#,##0"
                h_cell.alignment = _Style.RIGHT
                hours_val = float(row["hours"])  # type: ignore[arg-type]
                if hours_val > 80:
                    h_cell.fill = _Style.RED_BG
                elif hours_val > 40:
                    h_cell.fill = _Style.YELLOW_BG
                r_idx += 1
            # Totals row
            total = ws.cell(
                row=r_idx, column=1,
                value=f"Σ ИТОГО ({len(abs_rows)} отсутствий)",
            )
            total.font = _Style.BOLD_FONT
            total.fill = _Style.TOTALS_FILL
            ws.merge_cells(
                start_row=r_idx, start_column=1, end_row=r_idx, end_column=6,
            )
            days_total = sum(int(r["days"]) for r in abs_rows)  # type: ignore[misc,arg-type]
            hours_total = sum(float(r["hours"]) for r in abs_rows)  # type: ignore[misc,arg-type]
            c = ws.cell(row=r_idx, column=7, value=days_total)
            c.font = _Style.BOLD_FONT
            c.fill = _Style.TOTALS_FILL
            c.alignment = _Style.RIGHT
            c.number_format = "#,##0"
            c = ws.cell(row=r_idx, column=8, value=round(hours_total, 1))
            c.font = _Style.BOLD_FONT
            c.fill = _Style.TOTALS_FILL
            c.alignment = _Style.RIGHT
            c.number_format = "#,##0"

            # Тонкая линия снизу каждой строки отсутствий (без totals)
            _stamp_row_borders(ws, [(abs_header_row + 1, r_idx - 1, 8)])

        # --- Column widths ---
        widths = [22, 14, 18, 14, 12, 12, 8, 12]
        for c_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        ws.freeze_panes = "A2"

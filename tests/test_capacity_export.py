"""Тесты экспорта capacity в xlsx."""

import io
import pytest
from openpyxl import load_workbook

from app.models import Employee
from app.services.export_service import ExportService


@pytest.fixture
def team_employees(db_session):
    rows = [
        Employee(id="e1", jira_account_id="a1", display_name="Иванов", is_active=True, team="Alpha"),
        Employee(id="e2", jira_account_id="a2", display_name="Петров", is_active=True, team="Alpha"),
        Employee(id="e3", jira_account_id="a3", display_name="Сидоров", is_active=True, team=None),
    ]
    db_session.add_all(rows)
    db_session.commit()
    return rows


def test_export_capacity_xlsx_structure(db_session, team_employees):
    svc = ExportService(db_session)
    blob = svc.export_capacity_xlsx(2026, 2)

    assert isinstance(blob, bytes)
    assert len(blob) > 100

    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == "Сотрудник"
    # Three months × 3 cols each + 3 totals = 12 cols → 13 including first
    assert len(header) == 13

    body_names = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    assert "Alpha" in body_names           # team row
    assert "Без команды" in body_names      # unassigned group
    assert "Иванов" in body_names
    assert "Петров" in body_names
    assert "Сидоров" in body_names

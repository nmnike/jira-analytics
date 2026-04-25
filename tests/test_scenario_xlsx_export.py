"""Tests for ScenarioXlsxExporter — 4-sheet «Бухгалтерия» layout."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import (
    BacklogItem, Employee, EmployeeTeam, MandatoryWorkType,
    PlanningScenario, ScenarioAllocation, Role,
)
from app.services.scenario_xlsx_export import ScenarioXlsxExporter


EXPECTED_SHEETS = ["Сводка", "Включено", "Не вошло", "Справочник"]


@pytest.fixture
def minimal_scenario(db_session):
    """Минимальный сценарий: команда, два сотрудника, две задачи (одна вошла, одна нет)."""
    db_session.add_all([
        Role(code="dev", label="Разработчик", color="#1890FF",
             is_active=True, counts_in_planning=True),
        Role(code="analyst", label="Аналитик", color="#722ED1",
             is_active=True, counts_in_planning=True),
        MandatoryWorkType(code="org", label="Орг. вопросы",
                          is_active=True, subtracts_from_pool=True),
    ])
    db_session.flush()

    dave = Employee(
        jira_account_id="d1", display_name="Dave", role="dev", is_active=True,
    )
    alice = Employee(
        jira_account_id="a1", display_name="Alice", role="analyst", is_active=True,
    )
    db_session.add_all([dave, alice])
    db_session.flush()
    db_session.add_all([
        EmployeeTeam(employee_id=dave.id, team="Alpha", is_primary=True),
        EmployeeTeam(employee_id=alice.id, team="Alpha", is_primary=True),
    ])

    item_in = BacklogItem(
        title="Build feature", priority=1,
        estimate_hours=80, estimate_dev_hours=80,
    )
    item_out = BacklogItem(
        title="Skipped feature", priority=5,
        estimate_hours=200, estimate_dev_hours=200,
    )
    db_session.add_all([item_in, item_out])
    db_session.flush()

    scenario = PlanningScenario(
        name="Q2 2026 Alpha Base", year=2026, quarter="Q2",
        team="Alpha", status="draft",
    )
    db_session.add(scenario)
    db_session.flush()

    db_session.add_all([
        ScenarioAllocation(
            scenario_id=scenario.id, backlog_item_id=item_in.id,
            included_flag=True, planned_hours=80.0,
        ),
        ScenarioAllocation(
            scenario_id=scenario.id, backlog_item_id=item_out.id,
            included_flag=False, planned_hours=0.0,
        ),
    ])
    db_session.flush()

    class _R:
        pass
    r = _R()
    r.scenario_id = scenario.id
    return r


class TestSummaryHeader:
    def test_title_strip_in_a1(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        a1 = ws.cell(row=1, column=1).value or ""
        assert "Q2 2026 Alpha Base" in a1
        assert "черновик" in a1

    def test_subtitle_in_a2(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        a2 = ws.cell(row=2, column=1).value or ""
        assert "Q2" in a2
        assert "2026" in a2
        assert "Alpha" in a2
        assert "сформировано" in a2.lower()

    def test_summary_section_header_at_a4(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        assert ws.cell(row=4, column=1).value == "СВОДКА"

    def test_summary_metrics_table(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        # Row 5 labels
        assert ws.cell(row=5, column=1).value == "Ёмкость, ч"
        assert ws.cell(row=5, column=3).value == "План, ч"
        assert ws.cell(row=5, column=5).value == "Остаток, ч"
        assert ws.cell(row=5, column=7).value == "Использование"
        # Row 6 labels
        assert ws.cell(row=6, column=1).value == "Включено, шт."
        assert ws.cell(row=6, column=3).value == "Не вошло, шт."
        assert ws.cell(row=6, column=5).value == "QA дефицит, ч"
        assert ws.cell(row=6, column=7).value == "Отсутствия, ч"
        # Counts (row 6 col 2 — included = 1, col 4 — excluded = 1)
        assert ws.cell(row=6, column=2).value == 1
        assert ws.cell(row=6, column=4).value == 1


class TestSummaryByRoleMonths:
    def test_section_header_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any(v and "ПО РОЛЯМ" in str(v) and "МЕСЯЦАМ" in str(v) for v in labels)

    def test_role_table_headers_for_q2(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Роль":
                header_row = r
                break
        assert header_row is not None
        headers = [ws.cell(row=header_row, column=c).value for c in range(1, 9)]
        assert headers == ["Роль", "Апр", "Май", "Июн", "Σ Ёмкость", "План", "Остаток", "% исп."]

    def test_role_row_dev_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        found = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Разработчик":
                found = True
                assert ws.cell(row=r, column=6).value == pytest.approx(80.0)
                break
        assert found

    def test_totals_row_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "ИТОГО" in [str(v) for v in labels if v is not None]


class TestSummaryByEmployee:
    def test_section_header_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        labels = [str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)]
        assert any("ПО СОТРУДНИКАМ" in lbl for lbl in labels)

    def test_employee_table_headers(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        emp_header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Сотрудник":
                emp_header_row = r
                break
        assert emp_header_row is not None
        headers = [ws.cell(row=emp_header_row, column=c).value for c in range(1, 7)]
        assert headers == [
            "Сотрудник", "Роль", "Норма, ч", "Отсутствия, ч",
            "Доступно, ч", "Дней отсутствия",
        ]

    def test_dave_row_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]
        found = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Dave":
                found = True
                assert ws.cell(row=r, column=2).value == "Разработчик"
                assert (ws.cell(row=r, column=5).value or 0) > 0
                break
        assert found


INCLUDED_HEADERS_MID = [
    "Ключ Jira", "Название", "Приоритет", "Заказчик",
    "Аналитик, ч", "Разработка, ч", "QA, ч", "ОПЭ, ч",
    "Итого, ч", "План, ч", "Цели",
]


class TestIncludedSheet:
    def test_title_strip(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        a1 = ws.cell(row=1, column=1).value or ""
        assert "Q2 2026 Alpha Base" in a1
        assert "Включено" in a1
        assert "1" in a1  # the count "(1 задач)"

    def test_headers_in_row_2(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        header = [ws.cell(row=2, column=c).value for c in range(1, 12)]
        assert header == INCLUDED_HEADERS_MID

    def test_data_row_for_build_feature(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        assert ws.cell(row=3, column=2).value == "Build feature"
        assert ws.cell(row=3, column=3).value == 1
        assert ws.cell(row=3, column=6).value == pytest.approx(80.0)
        assert ws.cell(row=3, column=9).value == pytest.approx(80.0)  # Итого
        assert ws.cell(row=3, column=10).value == pytest.approx(80.0)  # План

    def test_totals_row_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        last = ws.max_row
        assert "ИТОГО" in str(ws.cell(row=last, column=1).value or "")
        assert ws.cell(row=last, column=9).value == pytest.approx(80.0)

    def test_autofilter_set(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        assert ws.auto_filter.ref is not None

    def test_freeze_panes_a3(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        assert ws.freeze_panes == "A3"


EXCLUDED_HEADERS_EXPECTED = [
    "Ключ Jira", "Название", "Приоритет", "Заказчик",
    "Аналитик, ч", "Разработка, ч", "QA, ч", "ОПЭ, ч",
    "Итого, ч", "Цели",
]


class TestExcludedSheet:
    def test_title_strip(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Не вошло"]
        a1 = ws.cell(row=1, column=1).value or ""
        assert "Не вошло" in a1
        assert "1" in a1

    def test_headers_no_plan_column(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Не вошло"]
        header = [ws.cell(row=2, column=c).value for c in range(1, 11)]
        assert header == EXCLUDED_HEADERS_EXPECTED

    def test_excluded_row_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Не вошло"]
        assert ws.cell(row=3, column=2).value == "Skipped feature"

    def test_data_rows_have_grey_fill(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Не вошло"]
        cell = ws.cell(row=3, column=2)
        assert cell.fill.fgColor.value.upper().endswith("FAFAFA")


class TestReferenceSheet:
    def test_three_section_headers_present(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Справочник"]
        labels = [str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)]
        joined = " ".join(labels)
        assert "ПРАВИЛА" in joined
        assert "ВНЕШНИЙ QA" in joined
        assert "ОТСУТСТВИЯ" in joined

    def test_external_qa_when_set(self, db_session, minimal_scenario):
        scenario = db_session.get(PlanningScenario, minimal_scenario.scenario_id)
        scenario.external_qa_hours = 120.0
        db_session.flush()

        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Справочник"]
        qa_value_seen = False
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v == 120 or v == 120.0 or (isinstance(v, str) and "120" in v):
                qa_value_seen = True
        assert qa_value_seen

    def test_external_qa_when_not_set(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Справочник"]
        all_text = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1) for c in range(1, 9)
        )
        assert "не задан" in all_text

    def test_no_absences_message(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Справочник"]
        all_text = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1) for c in range(1, 9)
        )
        assert "Отсутствий в квартале нет" in all_text

    def test_absences_table_when_present(self, db_session, minimal_scenario):
        from datetime import date as _d
        from app.models import Absence, AbsenceReason
        reason = AbsenceReason(
            code="vac", label="Отпуск", is_planned=True, color="#16A34A",
            sort_order=0,
        )
        db_session.add(reason)
        db_session.flush()
        emp = db_session.query(Employee).filter(Employee.display_name == "Dave").first()
        db_session.add(Absence(
            employee_id=emp.id,
            start_date=_d(2026, 4, 6),
            end_date=_d(2026, 4, 12),
            reason_id=reason.id,
            hours_total=40.0,
        ))
        db_session.flush()

        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Справочник"]

        found = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Dave":
                found = True
                assert ws.cell(row=r, column=3).value == "Отпуск"
        assert found


class TestStructure:
    def test_workbook_has_four_sheets_in_order(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        assert wb.sheetnames == EXPECTED_SHEETS

    def test_unknown_scenario_raises(self, db_session):
        with pytest.raises(ValueError, match="not found"):
            ScenarioXlsxExporter(db_session, "no-such-id").build()


@pytest.fixture
def empty_scenario(db_session):
    """Сценарий-пустышка."""
    scenario = PlanningScenario(
        name="Empty", year=2026, quarter="Q3", team="Ghost", status="draft",
    )
    db_session.add(scenario)
    db_session.flush()

    class _R:
        pass
    r = _R()
    r.scenario_id = scenario.id
    return r


class TestPlannedHoursAssigneeByJiraName:
    """Если задача не привязана к сотруднику через assignee_employee_id,
    но в Jira-issue указан исполнитель по имени — роль резолвится через
    лookup по display_name (как в /planning endpoint)."""

    def test_jira_assignee_name_routes_analyst_hours_to_pm(self, db_session):
        from app.models import Issue, Project
        db_session.add_all([
            Role(code="analyst", label="Аналитик", color="#722ED1",
                 is_active=True, counts_in_planning=True),
            Role(code="project_manager", label="Руководитель проектов",
                 color="#FA8C16", is_active=True, counts_in_planning=True),
        ])
        db_session.flush()

        analyst_emp = Employee(
            jira_account_id="anl2", display_name="Анна А.", role="analyst",
            is_active=True,
        )
        pm_emp = Employee(
            jira_account_id="pm2", display_name="Пётр П.", role="project_manager",
            is_active=True,
        )
        db_session.add_all([analyst_emp, pm_emp])
        db_session.flush()
        db_session.add_all([
            EmployeeTeam(employee_id=analyst_emp.id, team="Gamma", is_primary=True),
            EmployeeTeam(employee_id=pm_emp.id, team="Gamma", is_primary=True),
        ])

        # Issue с assignee_display_name = "Пётр П." (по имени, без linked employee)
        proj = Project(jira_project_id="px", key="GAMMA", name="Gamma project")
        db_session.add(proj)
        db_session.flush()
        issue = Issue(
            jira_issue_id="i100", key="GAMMA-1",
            summary="Issue with PM name", issue_type="Task", status="Open",
            project_id=proj.id,
            assignee_display_name="Пётр П.",
        )
        db_session.add(issue)
        db_session.flush()

        # Backlog item — ассоциация только через issue, без assignee_employee_id
        item = BacklogItem(
            title="PM-driven via Jira",
            priority=1,
            estimate_hours=100,
            estimate_analyst_hours=100,
            issue_id=issue.id,
            # assignee_employee_id NOT set
        )
        db_session.add(item)
        db_session.flush()

        scenario = PlanningScenario(
            name="Gamma plan", year=2026, quarter="Q2",
            team="Gamma", status="draft",
        )
        db_session.add(scenario)
        db_session.flush()
        db_session.add(ScenarioAllocation(
            scenario_id=scenario.id, backlog_item_id=item.id,
            included_flag=True, planned_hours=100.0,
        ))
        db_session.flush()

        data = ScenarioXlsxExporter(db_session, scenario.id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]

        analyst_plan = None
        pm_plan = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label == "Аналитик":
                analyst_plan = ws.cell(row=r, column=6).value
            elif label == "Руководитель проектов":
                pm_plan = ws.cell(row=r, column=6).value
        assert analyst_plan in (0, 0.0, None)
        assert pm_plan == pytest.approx(100.0)


class TestPlannedHoursAssigneeSubstitution:
    """Аналитический объём задачи, назначенной на РП/Консультанта,
    должен попасть в его роль, а не в «Аналитик»."""

    def test_pm_assignee_routes_analyst_hours_to_pm(self, db_session):
        # Setup: команда, аналитик, РП. Одна задача с назначенным РП.
        db_session.add_all([
            Role(code="analyst", label="Аналитик", color="#722ED1",
                 is_active=True, counts_in_planning=True),
            Role(code="project_manager", label="Руководитель проектов",
                 color="#FA8C16", is_active=True, counts_in_planning=True),
        ])
        db_session.flush()

        analyst_emp = Employee(
            jira_account_id="anl", display_name="Анна А.", role="analyst",
            is_active=True,
        )
        pm_emp = Employee(
            jira_account_id="pm1", display_name="Пётр П.", role="project_manager",
            is_active=True,
        )
        db_session.add_all([analyst_emp, pm_emp])
        db_session.flush()
        db_session.add_all([
            EmployeeTeam(employee_id=analyst_emp.id, team="Beta", is_primary=True),
            EmployeeTeam(employee_id=pm_emp.id, team="Beta", is_primary=True),
        ])

        item = BacklogItem(
            title="PM-driven feature",
            priority=1,
            estimate_hours=100,
            estimate_analyst_hours=100,
            assignee_employee_id=pm_emp.id,
        )
        db_session.add(item)
        db_session.flush()

        scenario = PlanningScenario(
            name="Beta plan", year=2026, quarter="Q2",
            team="Beta", status="draft",
        )
        db_session.add(scenario)
        db_session.flush()
        db_session.add(ScenarioAllocation(
            scenario_id=scenario.id, backlog_item_id=item.id,
            included_flag=True, planned_hours=100.0,
        ))
        db_session.flush()

        data = ScenarioXlsxExporter(db_session, scenario.id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Сводка"]

        # На «По ролям × месяцам»: ищем строку РП и Аналитик, проверяем «План»
        analyst_plan = None
        pm_plan = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label == "Аналитик":
                analyst_plan = ws.cell(row=r, column=6).value
            elif label == "Руководитель проектов":
                pm_plan = ws.cell(row=r, column=6).value
        # Аналитик должен быть 0 (исполнитель — РП), РП — 100
        assert analyst_plan in (0, 0.0, None)
        assert pm_plan == pytest.approx(100.0)


class TestWrapAndBorders:
    """wrap_text=True и тонкая нижняя граница на data-строках."""

    def test_data_cell_has_wrap_text(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        # Cell with the title "Build feature"
        cell = ws.cell(row=3, column=2)
        assert cell.alignment.wrap_text is True

    def test_data_row_has_bottom_border(self, db_session, minimal_scenario):
        data = ScenarioXlsxExporter(db_session, minimal_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        ws = wb["Включено"]
        cell = ws.cell(row=3, column=2)
        assert cell.border.bottom is not None
        assert cell.border.bottom.style == "thin"


class TestEmpty:
    def test_build_does_not_crash(self, db_session, empty_scenario):
        data = ScenarioXlsxExporter(db_session, empty_scenario.scenario_id).build()
        assert data[:2] == b"PK"
        wb = load_workbook(BytesIO(data))
        assert wb.sheetnames == EXPECTED_SHEETS

    def test_empty_sections_render_safely(self, db_session, empty_scenario):
        data = ScenarioXlsxExporter(db_session, empty_scenario.scenario_id).build()
        wb = load_workbook(BytesIO(data))
        # Сводка — title strip + section headers should still be there
        ws = wb["Сводка"]
        assert ws.cell(row=4, column=1).value == "СВОДКА"
        # Справочник — "Отсутствий в квартале нет" message
        ws_ref = wb["Справочник"]
        all_text = " ".join(
            str(ws_ref.cell(row=r, column=c).value or "")
            for r in range(1, ws_ref.max_row + 1) for c in range(1, 9)
        )
        assert "Отсутствий в квартале нет" in all_text
